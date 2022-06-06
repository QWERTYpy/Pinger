import openpyxl  # Библиотека для работы с таблицами
from init_const import dict_const
import time
from ping3 import ping


class ListsIP:
    def __init__(self):
        ip_book = openpyxl.load_workbook(dict_const['name_file'])  # Открывает файл
        self.worksheet = ip_book.active  # Делаем его активным
        self.max_row = self.worksheet.max_row  # Получаем максимальное количество строк
        self.max_col = self.worksheet.max_column  # Получаем максимальное количество столбцов

        # Инициализация начальных данных
        self.scan_ip = []  # Список всех устройств
        self.off_ip = []  # Список отсутствующих
        self.reboot_ping = False
        self.excel_to_list()

    def excel_to_list(self):
        """
        Функция заполняет список устройствами для опроса

        :return:
        Список списков устройств
        """
        scan_ip = []
        for row in range(1, self.max_row + 1):  # Запускаем цикл по всем строкам
            ip_type = self.worksheet.cell(row=row, column=dict_const['ip_type']).value  # Тип устройства
            if ip_type in dict_const['type_name']:  # Если устройство есть в списке
                ip_cam = self.worksheet.cell(row=row, column=dict_const['ip_camera']).value  # Адрес устройства
                ip_object = self.worksheet.cell(row=row,
                                           column=dict_const['ip_object']).value  # Объект на котором оно установлена
                ip_comment = self.worksheet.cell(row=row, column=dict_const['ip_comment']).value  # Комментарий
                ip_priority = self.worksheet.cell(row=row,
                                             column=dict_const['ip_priority']).value  # Важность для отображения
                ip_active = self.worksheet.cell(row=row,
                                           column=dict_const['ip_active']).value  # Работоспособность устройства
                if ip_priority != dict_const['not_important']:  # Если устройство не исключено из наблюдения
                    self.scan_ip.append([0, row, ip_cam, ip_object, ip_type, ip_comment, ip_priority, ip_active])

    def modification_off_ip(self, flag_ip, row, ip_cam, ip_object, ip_type, ip_comment):
        """
        Данная функция изменяет список с отсутствующими камерами, добавляя новые или помечая на удаление старые

        :param flag_ip: bool Ответило устройство на пинг или нет
        :param row: Строка в таблице с устройством
        :param ip_cam: IP устройства
        :param ip_object: Объект на котором установлено устройство
        :param ip_type: Тип устройства
        :param ip_comment: Комментарий
        :return: Время остановки устройства
        """
        flag_add = True  # По умолчанию новую запись помечаем на добавление
        time_end = 0
        for i in range(len(self.off_ip)):
            if ip_cam == self.off_ip[i][2]:  # Если устройство присутствует в списке
                flag_add = False  # Помечаем, что добавлять его не надо
                if flag_ip:
                    self.off_ip[i][0] = 2  # Если ответило на пинг, помечаем на удаление из списка
                else:
                    self.off_ip[i][0] = 1  # Если оно снова не ответило на пинг отключаем уведомление
                time_end = self.off_ip[i][6]

        if flag_add and not flag_ip:  # Если устройства нет в списке и оно не отвечает - добавляем
            time_end = int(time.time())
            self.off_ip.append([0, row, ip_cam, ip_object, ip_type, ip_comment, time_end])
        return time_end

    @staticmethod
    def thread_ping(list_ip):
        if list_ip[7] == 'ON':
            ip_pin = ping(list_ip[2])
            list_ip[0] = 1  # Пинг прошёл
            if ip_pin is None or type(ip_pin) is not float:
                ip_pin = ping(list_ip[2])
                if ip_pin is None or type(ip_pin) is not float:
                    list_ip[0] = 0  # Пинг не прошёл

    def hidden_ping(self, off_ip):
        """
        Процедура которая пингует отсутсвующие устройства через 5 сек, для подтвержения отсуствия связи
        :return:
        """
        time.sleep(5)
        for flag, _, ip, obj, _type, com, tm in off_ip:
            ip_pin = ping(ip)
            flag_ip = True  # Пинг прошёл
            if ip_pin is None or type(ip_pin) is not float:
                flag_ip = False
            if flag_ip:
                self.reboot_ping = True
                break
            if self.reboot_ping:
                break

    def message(self):
        self.reboot_ping = False
        off_text = ''
        on_text = ''
        del_index = []
        file_log = open(dict_const['log_file'], 'a')
        for flag, _, ip, obj, _type, com, tm in self.off_ip:
            if not flag:  # Если устройство впервые не ответило, выводим сообщение и пишем в лог
                off_text += f'{ip[10:]},'
                file_log.write(f'OFF >> {ip} - {time.ctime()}\n')
            if flag == 2:  # Если устройство ответило, выводим сообщение, пишем в лог, удаляем из списка
                on_text += f'{ip[10:]},'
                file_log.write(f'ONN >> {ip} - {time.ctime()}\n')
                del_index.append(self.off_ip.index([flag, _, ip, obj, _type, com, tm]))
        file_log.close()
        # Удаляем появившиеся на связи устройства
        del_index.reverse()
        for _ in del_index:
            self.off_ip.pop(_)
        return on_text, off_text
