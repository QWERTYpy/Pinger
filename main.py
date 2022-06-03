"""
Скрипт для опроса сетевых устройств
"""
import sys
import openpyxl  # Библиотека для работы с таблицами
from ping3 import ping
import win10toast  # Библиотека для всплывающих сообщений
import time
import os
import msvcrt
from rich.console import Console
from rich.table import Table
from rich.progress import track
import datetime
from collections import deque
from init_const import dict_const
import threading
import concurrent.futures


ip_book = openpyxl.load_workbook(dict_const['name_file'])  # Открывает файл
worksheet = ip_book.active  # Делаем его активным
max_row = worksheet.max_row  # Получаем максимальное количество строк
max_col = worksheet.max_column  # Получаем максимальное количество столбцов

# Инициализация начальных данных
scan_ip = []  # Список всех устройств
off_ip = []  # Список отсутствующих
reboot_ping = False
console = Console()


def table_generation(title_table, columns_table, _type='Main'):
    """ Функция создает таблицы для форматирования вывода консоли

    :param title_table: Заголовок таблицы
    :param columns_table: Список заголовков столбцов
    :param _type: Видимая таблица или для зонирования
    :return:
    Указатель на созданную таблицу
    """
    if _type == 'Main':
        table = Table(title=f'{title_table}')
    elif _type == 'Grid':
        table = Table.grid(expand=True)
    for column in columns_table:
        if isinstance(column, str):
            table.add_column(column)
        if isinstance(column, list):
            table.add_column(column[0], width=column[1])

    return table


def excel_to_list():
    """
    Функция заполняет список устройствами для опроса

    :return:
    Список списков устройств
    """
    scan_ip = []
    for row in range(1, max_row+1):  # Запускаем цикл по всем строкам
        ip_type = worksheet.cell(row=row, column=dict_const['ip_type']).value  # Тип устройства
        if ip_type in dict_const['type_name']:  # Если устройство есть в списке
            ip_cam = worksheet.cell(row=row, column=dict_const['ip_camera']).value  # Адрес устройства
            ip_object = worksheet.cell(row=row, column=dict_const['ip_object']).value  # Объект на котором оно установлена
            ip_comment = worksheet.cell(row=row, column=dict_const['ip_comment']).value  # Комментарий
            ip_priority = worksheet.cell(row=row, column=dict_const['ip_priority']).value  # Важность для отображения
            ip_active = worksheet.cell(row=row, column=dict_const['ip_active']).value  # Работоспособность устройства
            if ip_priority != dict_const['not_important']:  # Если устройство не исключено из наблюдения
                scan_ip.append([0, row, ip_cam, ip_object, ip_type, ip_comment, ip_priority, ip_active])
    return scan_ip


def modification_off_ip(flag_ip, row, ip_cam, ip_object, ip_type, ip_comment):
    """
    Данная функция изменяет список с отсутствующими камерами, добавляя новые или помечая на удаление старые

    :param flag_ip: bool Ответило устройство на пинг или нет
    :param row: Строка в таблице с устройством
    :param ip_cam: IP устройства
    :param ip_object: Объект на котором установлено устройство
    :param ip_type: Тип устройства
    :param ip_comment: Комментарий
    :return: Время остановки устройства
    """
    flag_add = True  # По умолчанию новую запись помечаем на добавление
    time_end = 0
    for i in range(len(off_ip)):
        if ip_cam == off_ip[i][2]:  # Если устройство присутствует в списке
            flag_add = False  # Помечаем, что добавлять его не надо
            if flag_ip:
                off_ip[i][0] = 2  # Если ответило на пинг, помечаем на удаление из списка
            else:
                off_ip[i][0] = 1  # Если оно снова не ответило на пинг отключаем уведомление
            time_end = off_ip[i][6]

    if flag_add and not flag_ip:  # Если устройства нет в списке и оно не отвечает - добавляем
        time_end = int(time.time())
        off_ip.append([0, row, ip_cam, ip_object, ip_type, ip_comment, time_end])
    return time_end


def print_time_input(timeout):
    """
    Данная процедура выводит обратный таймер и следит за вводом для управления программой

    :param timeout: Значение таймера

    :return:
    """
    start = time.monotonic()
    now = time.monotonic()
    count_sec = 0
    global reboot_ping
    while now - start < timeout:
        if now - start > count_sec:
            count_sec += 1
            print_time = timeout-count_sec
            print("\rОсталось %03d c. | u-обновить | t-таймер | q-выход :" % print_time, end='')
        if msvcrt.kbhit():
            c = msvcrt.getwch()
            if c == 'q' or c == 'й':
                reboot_ping = True
                sys.exit()
            if c == 'u' or c == 'г':
                break
            if c == 't' or c == 'е':
                dict_const['time_sec'] = int(input('Введите количество секунд :'))
                break
        if reboot_ping:
            break
        now = time.monotonic()


def thread_ping(list_ip):
    if list_ip[7] == 'ON':
        ip_pin = ping(list_ip[2])
        list_ip[0] = 1  # Пинг прошёл
        if ip_pin is None or type(ip_pin) is not float:
            ip_pin = ping(list_ip[2])
            if ip_pin is None or type(ip_pin) is not float:
                list_ip[0] = 0  # Пинг не прошёл


def tab_ping(scan_ip):
    """
    Процедура отвечающая за отрисовку консоли согласно полученных данных
    :return:
    """
    count_device = len(scan_ip)
    #  Создаем необходимые таблицы для визуального оформления
    table_priority = table_generation(dict_const['table_priority_name'], dict_const['table_priority_field'])
    table_ip = table_generation(dict_const['table_ip_name'], dict_const['table_ip_field'])
    table_error = table_generation(dict_const['table_error_name'], dict_const['table_error_field'])
    grid_main = table_generation("", dict_const['console'][2:4], 'Grid')
    grid_left = table_generation("", ("",), 'Grid')
    grid_left.add_row(table_priority)
    grid_left.add_row(table_error)
    # Создаем таблицу для вывода логов
    table_log = Table(title="История опросов")
    table_log.add_column('IP Адрес - Время')
    grid_main.add_row(grid_left, table_log)

    # Создаем пул на 200 потоков
    with concurrent.futures.ThreadPoolExecutor(max_workers=200) as executor:
        executor.map(thread_ping, scan_ip)

    for flag_ip, row, ip_cam, ip_object, ip_type, ip_comment, ip_priority, ip_active in scan_ip:
        if ip_active == 'ON':  # Если устройство включено
            # Изменяем список устройств не отвечающих на пинг
            time_end = modification_off_ip(flag_ip, row, ip_cam, ip_object, ip_type, ip_comment)
            if ip_priority == dict_const['important']:  # Заполняем таблицу приоритетных устройств
                if flag_ip:
                    table_priority.add_row(f'[green]{ip_cam}[/green]', str(True if flag_ip else False), ip_object, ip_type, ip_comment)
                else:
                    table_priority.add_row(f'[red]{ip_cam}[/red]', str(True if flag_ip else False), ip_object, ip_type, ip_comment)
            if not flag_ip:  # Заполняем таблицу не ответивших устройств
                time_format = str(datetime.timedelta(seconds=int(time.time())-time_end))
                table_ip.add_row(f'[red]{ip_cam}[/red]', str(True if flag_ip else False), ip_object, ip_type, ip_comment, time_format)
        if ip_active == 'OFF':  # Если устройства отключены заполняем таблицу отключенных
            table_error.add_row(f'[red]{ip_cam}[/red]', ip_object, ip_type, ip_comment)
    #  Заполняем строки таблицами
    grid_left.add_row(f'Всего устройств: {str(count_device)} '
                      f'[green] На связи: {count_device-table_ip.row_count-table_error.row_count}[/green]'
                      f'[red] Отсутствуют: {table_ip.row_count}[/red]'
                      f'[blue] Отключены: {table_error.row_count}[/blue]')
    if table_ip.row_count > 0:  # Если таблица не пустая - выводим
        grid_left.add_row(table_ip)

    off_text = ''
    on_text = ''
    del_index = []
    file_log = open(dict_const['log_file'], 'a')
    for flag, _, ip, obj, _type, com, tm in off_ip:
        if not flag:  # Если устройство впервые не ответило, выводим сообщение и пишем в лог
            off_text += f'{ip[10:]},'
            file_log.write(f'OFF >> {ip} - {time.ctime()}\n')
        if flag == 2:  # Если устройство ответило, выводим сообщение, пишем в лог, удаляем из списка
            on_text += f'{ip[10:]},'
            file_log.write(f'ONN >> {ip} - {time.ctime()}\n')
            del_index.append(off_ip.index([flag, _, ip, obj, _type, com, tm]))
    file_log.close()
    # Удаляем появившиеся на связи устройства
    del_index.reverse()
    for _ in del_index:
        off_ip.pop(_)
    # Выводим всплывающее собщение
    toaster = win10toast.ToastNotifier()
    if len(off_text) > 0:
        toaster.show_toast("Отсутствуют", off_text[:-1], icon_path='ping.ico', duration=5, threaded=True)
    if len(on_text) > 0:
        toaster.show_toast("Восстановление связи", on_text[:-1], icon_path='ping.ico', duration=5, threaded=True)
    # Заполняем таблицу событий
    file_log = open(dict_const['log_file'], 'r')
    my_stack = deque(maxlen=dict_const['buffer'])  # Выводим согласно размера буфера
    for line_log in file_log:
        my_stack.append(line_log[:-1])
    file_log.close()
    for stack_line in my_stack:
        if stack_line[0:3] == 'ONN':
            table_log.add_row(f'[green]{stack_line}[/green]')
        if stack_line[0:3] == 'OFF':
            table_log.add_row(f'[red]{stack_line}[/red]')
    # Выводим таблицы на экран
    os.system(f"mode con:cols={dict_const['console'][0]} lines={dict_const['console'][1]}")
    console.print(grid_main)


def hidden_ping():
    """
    Процедура которая пингует отсутсвующие устройства через 5 сек, для подтвержения отсуствия связи
    :return:
    """
    time.sleep(5)
    global reboot_ping
    for flag, _, ip, obj, _type, com, tm in off_ip:
        ip_pin = ping(ip)
        flag_ip = True  # Пинг прошёл
        if ip_pin is None or type(ip_pin) is not float:
            flag_ip = False
        if flag_ip:
            reboot_ping = True
            break
        if reboot_ping:
            break


while True:
    if len(scan_ip) == 0:
        scan_ip = excel_to_list()
    tab_ping(scan_ip)
    reboot_ping = False
    if len(off_ip) > 0:
        threading.Thread(target=hidden_ping).start()
    print_time_input(dict_const['time_sec'])
    os.system('cls')
